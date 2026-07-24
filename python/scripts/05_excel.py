#!/usr/bin/env python3
"""
Paso 5 (opcional) — Exportar las tablas del reporte a un Excel con formato.

Lee los tres CSV que dejó 04_reporte.py y produce UN archivo .xlsx con tres
pestañas, ya presentable para estudiar: encabezados con color, columnas anchas,
números alineados con decimales razonables, filas alternadas, y resaltado de
los coeficientes que cambiaron de signo.

Uso:
    python python/scripts/05_excel.py --dataset swissmetro

Requiere openpyxl (se instala con: pip install openpyxl).
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from cvae.config_repo import cargar  # noqa: E402

# --- paleta -------------------------------------------------------------- #
AZUL = "1F3A5C"
AZUL_CLARO = "DCE6F1"
GRIS_FILA = "F5F7FA"
ROJO_SUAVE = "F8D7DA"
VERDE_SUAVE = "D5E8D8"
BLANCO = "FFFFFF"

FUENTE = "Arial"

borde_fino = Border(*(Side(style="thin", color="C8CED6"),) * 4)


def _encabezado(celda):
    celda.font = Font(name=FUENTE, bold=True, color=BLANCO, size=11)
    celda.fill = PatternFill("solid", fgColor=AZUL)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = borde_fino


def _volcar(ws, df: pd.DataFrame, incluir_indice: bool):
    """Escribe un DataFrame en la hoja y aplica formato base."""
def _volcar(ws, df: pd.DataFrame, incluir_indice: bool):
    """Escribe un DataFrame en la hoja y aplica formato base."""
    # Si el índice tiene valor (p.ej. 'metodo'), lo convertimos en una columna
    # normal. Así evitamos la fila fantasma que dataframe_to_rows inserta bajo
    # el encabezado cuando se escribe con index=True.
    if incluir_indice:
        df = df.reset_index()

    filas = dataframe_to_rows(df, index=False, header=True)
    for r, fila in enumerate(filas, start=1):
        for c, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=c, value=valor)
            if r == 1:
                _encabezado(celda)
            else:
                celda.font = Font(name=FUENTE, size=10)
                celda.border = borde_fino
                if isinstance(valor, float):
                    celda.number_format = "0.0000"
                    celda.alignment = Alignment(horizontal="right")
                else:
                    celda.alignment = Alignment(horizontal="left")


def _rayado(ws, primera_datos: int):
    """Filas alternadas para lectura cómoda."""
    for r in range(primera_datos, ws.max_row + 1):
        if (r - primera_datos) % 2 == 1:
            for c in range(1, ws.max_column + 1):
                celda = ws.cell(row=r, column=c)
                if celda.fill.fgColor.rgb in (None, "00000000"):
                    celda.fill = PatternFill("solid", fgColor=GRIS_FILA)


def _anchos(ws):
    """Ajusta el ancho de columna al contenido."""
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        largo = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letra].width = min(max(largo + 3, 10), 30)
    ws.freeze_panes = "A2"


def exportar(dataset: str) -> Path:
    cfg = cargar()
    res = cfg.rutas.results

    fuentes = {
        "Predictiva": (res / f"tabla_predictiva_{dataset}.csv", True),
        "Coeficientes": (res / f"tabla_coeficientes_{dataset}.csv", False),
        "Cambios de signo": (res / f"tabla_signos_{dataset}.csv", False),
    }
    faltan = [str(p) for _, (p, _) in fuentes.items() if not p.exists()]
    if faltan:
        raise SystemExit(
            "Faltan archivos del paso 4:\n  " + "\n  ".join(faltan) +
            "\nCorre primero: python python/scripts/04_reporte.py --dataset " + dataset)

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    for nombre_hoja, (ruta, con_indice) in fuentes.items():
        try:
            df = pd.read_csv(ruta, index_col=0 if con_indice else None)
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()

        if df.empty:
            # Una tabla vacía casi siempre significa que falta algo en el
            # config: por ejemplo, `signos_esperados` sin entradas, o un
            # bloque `atributos` vacío (el modelo quedaría solo con constantes).
            ws = wb.create_sheet(nombre_hoja)
            ws["A1"] = f"Sin datos: {ruta.name} está vacío."
            ws["A2"] = "Revisa el bloque de este dataset en config.yaml."
            ws["A1"].font = Font(name=FUENTE, bold=True, size=11)
            ws["A2"].font = Font(name=FUENTE, size=10)
            ws.column_dimensions["A"].width = 60
            print(f"  AVISO: '{nombre_hoja}' quedó vacía ({ruta.name}).")
            continue

        ws = wb.create_sheet(nombre_hoja)
        _volcar(ws, df, incluir_indice=con_indice)
        _rayado(ws, primera_datos=2)
        _anchos(ws)

        # Resaltado específico de la hoja de cambios de signo.
        if nombre_hoja == "Cambios de signo" and "invertido_en_media" in df.columns:
            col_idx = list(df.columns).index("invertido_en_media") + 1
            for r in range(2, ws.max_row + 1):
                invertido = ws.cell(row=r, column=col_idx).value
                relleno = ROJO_SUAVE if str(invertido).lower() in ("true", "1") else VERDE_SUAVE
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=relleno)

    destino = res / f"tablas_{dataset}.xlsx"
    wb.save(destino)
    return destino


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dataset", required=True)
    args = ap.parse_args()

    destino = exportar(args.dataset)
    print(f"Excel generado: {destino}")
    print("Tres pestañas: Predictiva, Coeficientes, Cambios de signo.")
    print("En 'Cambios de signo', rojo = coeficiente con signo invertido en media.")


if __name__ == "__main__":
    main()